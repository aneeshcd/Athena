from __future__ import annotations

from pathlib import Path

from fastapi import HTTPException
from fastapi.testclient import TestClient
from neo4j.exceptions import ServiceUnavailable
from openpyxl import Workbook

import pytest

from app.config import get_settings, Settings
from app.main import app
from app.plugins import neo4jGraphPlugin
from app.plugins.neo4jGraphPlugin import impactAnalysisLLM
from app.plugins.neo4jGraphPlugin.impactAnalysisLLM import (
    HITL_NOTICE,
    ImpactAnalysisFailed,
    ImpactAnalysisUnavailable,
    build_coverage_flags,
    build_impact_facts,
    generateImpactAnalysis,
)
from app.plugins.neo4jGraphPlugin.ingestExcel import read_excel_artefact
from app.plugins.neo4jGraphPlugin.graphRepository import _safe_depth
from app.plugins.neo4jGraphPlugin.ontologyValidator import sanitize_identifier, validate_edges
from app.plugins.neo4jGraphPlugin.types import (
    ImpactEdge,
    ImpactGraph,
    ImpactNode,
    LLMImpactAnalysisInput,
    RequirementCandidate,
    SelectedRequirement,
)


def test_excel_ingestion_reads_required_sheets(tmp_path: Path):
    workbook_path = tmp_path / "artefact.xlsx"
    workbook = Workbook()
    nodes = workbook.active
    nodes.title = "nodes"
    nodes.append(["id", "type", "name", "description", "criticality"])
    nodes.append(["REQ-008", "Requirement", "Battery backup", "Battery backup duration", "High"])
    nodes.append(["SUB-001", "Subsystem", "Power subsystem", "UPS and battery pack", "Medium"])
    edges = workbook.create_sheet("edges")
    edges.append(["source_id", "target_id", "relationship", "description"])
    edges.append(["REQ-008", "SUB-001", "ALLOCATED_TO", "Allocated to power"])
    ontology = workbook.create_sheet("ontology")
    ontology.append(["source_entity", "relationship", "target_entity"])
    ontology.append(["Requirement", "ALLOCATED_TO", "Subsystem"])
    workbook.save(workbook_path)

    artefact = read_excel_artefact(str(workbook_path))

    assert artefact.nodes[0].id == "REQ-008"
    assert artefact.nodes[0].type == "Requirement"
    assert artefact.edges[0].relationship == "ALLOCATED_TO"
    assert artefact.ontology[0].target_entity == "Subsystem"


def test_ontology_validation_reports_invalid_edges(tmp_path: Path):
    artefact = read_excel_artefact(str(_sample_workbook(tmp_path)))

    invalid_edges = validate_edges(artefact.nodes, artefact.edges, artefact.ontology)

    assert [edge.relationship for edge in invalid_edges] == ["OWNED_BY"]
    assert invalid_edges[0].source_type == "Requirement"
    assert invalid_edges[0].target_type == "Team"


def test_requirement_search_contract_with_fake_repository():
    repository = FakeRepository()

    matches = repository.search_requirement("Adjust the backup power duration")

    assert matches[0].id == "REQ-008"
    assert matches[0].score > matches[1].score


def test_impact_traversal_contract_with_fake_repository():
    repository = FakeRepository()

    graph = repository.get_impact_graph("REQ-008", depth=2)

    assert {node.id: node.status for node in graph.nodes}["REQ-008"] == "selected"
    assert {"SUB-001", "TEST-004", "TEAM-POWER", "RISK-002"}.issubset({node.id for node in graph.nodes})
    assert all(edge.status == "ontology-link" for edge in graph.edges)


def test_change_analysis_returns_selected_requirement_and_impact_graph(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    result = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration")

    assert result.selectedRequirement.id == "REQ-008"
    assert result.selectedRequirement.type == "Requirement"
    assert not hasattr(result.selectedRequirement, "score")
    assert {node.id: node.status for node in result.impactGraph.nodes}["REQ-008"] == "selected"
    assert result.impactGraph.edges


def test_repeated_same_request_returns_identical_result(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    first = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration", requestId="same")
    second = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration", requestId="same")

    assert first.model_dump() == second.model_dump()


def test_different_requests_return_different_requirements_and_graphs(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    backup = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration", requestId="backup")
    flight = neo4jGraphPlugin.analyzeRequirementChange("Reduce the flight control response time", requestId="flight")

    assert backup.selectedRequirement.id == "REQ-008"
    assert flight.selectedRequirement.id == "REQ-004"
    assert backup.impactGraph.model_dump() != flight.impactGraph.model_dump()


def test_no_match_returns_empty_graph_and_clear_message(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    result = neo4jGraphPlugin.analyzeRequirementChange("nonsense unrelated text", requestId="bad")

    assert result.requestId == "bad"
    assert result.selectedRequirement is None
    assert result.impactGraph.nodes == []
    assert result.impactGraph.edges == []
    assert result.message == "No matching requirement found."


def test_debug_payload_excludes_search_candidates_from_impact_graph(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    result = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration", requestId="debug", debug=True)

    graph_node_ids = {node.id for node in result.impactGraph.nodes}
    excluded_ids = {candidate.id for candidate in result.debug.excludedCandidates}
    assert "REQ-002" in excluded_ids
    assert "REQ-002" not in graph_node_ids
    assert result.debug.traversalDepth == 2
    assert result.debug.nodeCount == len(result.impactGraph.nodes)
    assert result.debug.edgeCount == len(result.impactGraph.edges)


def test_duplicate_nodes_and_edges_are_not_returned_by_api_contract(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    result = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration")

    node_ids = [node.id for node in result.impactGraph.nodes]
    edge_ids = [edge.id for edge in result.impactGraph.edges]
    assert len(node_ids) == len(set(node_ids))
    assert len(edge_ids) == len(set(edge_ids))


def test_traversal_depth_is_clamped_to_supported_range(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    monkeypatch.setattr(neo4jGraphPlugin, "getImpactGraph", repository.get_impact_graph)

    result = neo4jGraphPlugin.analyzeRequirementChange("Adjust the backup power duration", depth=99, debug=True)

    assert _safe_depth(0) == 1
    assert _safe_depth(2) == 2
    assert _safe_depth(99) == 3
    assert result.debug.traversalDepth == 3


def test_requirement_search_endpoint_does_not_expose_scores(monkeypatch):
    repository = FakeRepository()
    monkeypatch.setattr(neo4jGraphPlugin, "searchRequirement", repository.search_requirement)
    client = TestClient(app)

    response = client.post("/api/graph/requirement-search", json={"changeText": "Adjust the backup power duration"})

    assert response.status_code == 200
    assert response.json()[0]["id"] == "REQ-008"
    assert "score" not in response.json()[0]


def test_required_graph_routes_are_registered():
    routes = {route.path for route in app.routes}

    assert {
        "/api/graph/ingest",
        "/api/graph/impact-from-change",
        "/api/graph/impact-analysis",
        "/api/graph/requirement-search",
        "/api/graph/impact",
        "/api/graph/ontology",
        "/api/graph/health",
        "/api/graph/llm-health",
        "/api/graph/llm-test",
    }.issubset(routes)


def test_neo4j_settings_default_to_compose_service_uri():
    settings = Settings(_env_file=None)

    assert settings.neo4j_uri == "bolt://neo4j:7687"


def test_neo4j_env_overrides_default_for_host_backend(monkeypatch):
    monkeypatch.setenv("NEO4J_URI", "bolt://localhost:7687")
    get_settings.cache_clear()

    settings = Settings(_env_file=None)

    assert settings.neo4j_uri == "bolt://localhost:7687"


def test_docker_without_neo4j_env_uses_compose_service_uri(monkeypatch):
    monkeypatch.delenv("NEO4J_URI", raising=False)
    monkeypatch.setenv("RUNNING_IN_DOCKER", "1")
    get_settings.cache_clear()

    settings = Settings(_env_file=None)

    assert settings.neo4j_uri == "bolt://neo4j:7687"


def test_docker_compose_uses_neo4j_environment_fallbacks():
    compose = Path(__file__).resolve().parents[2] / "docker-compose.yml"
    content = compose.read_text(encoding="utf-8")

    assert "NEO4J_URI: ${NEO4J_URI:-bolt://neo4j:7687}" in content
    assert "NEO4J_USERNAME: ${NEO4J_USERNAME:-neo4j}" in content
    assert "NEO4J_PASSWORD: ${NEO4J_PASSWORD}" in content


def test_env_example_documents_compose_neo4j_uri():
    env_example = Path(__file__).resolve().parents[2] / ".env.example"
    content = env_example.read_text(encoding="utf-8")

    assert "NEO4J_URI=bolt://neo4j:7687" in content
    assert "NEO4J_USERNAME=neo4j" in content
    assert "NEO4J_PASSWORD=<password>" in content


def test_dockerignore_excludes_python_caches_and_local_envs():
    root_ignore = Path(__file__).resolve().parents[2] / ".dockerignore"
    backend_ignore = Path(__file__).resolve().parents[1] / ".dockerignore"

    for ignore_file in (root_ignore, backend_ignore):
        content = ignore_file.read_text(encoding="utf-8")
        assert "__pycache__/" in content
        assert "*.pyc" in content
        assert ".venv/" in content
        assert "node_modules/" in content


def test_backend_dockerfile_disables_runtime_cache_copy():
    dockerfile = Path(__file__).resolve().parents[1] / "Dockerfile"
    content = dockerfile.read_text(encoding="utf-8")

    assert "PYTHONDONTWRITEBYTECODE=1" in content
    assert "RUNNING_IN_DOCKER=1" in content
    assert 'find /app -type d -name "__pycache__"' in content
    assert 'find /app -type f -name "*.pyc"' in content


def test_ingest_logs_resolved_neo4j_uri_before_ingesting(monkeypatch, caplog, tmp_path):
    from app.plugins.neo4jGraphPlugin import routes

    workbook_path = _sample_workbook(tmp_path)
    upload_bytes = workbook_path.read_bytes()

    class FakeUpload:
        filename = "artefact.xlsx"

        async def read(self):
            return upload_bytes

    monkeypatch.setenv("NEO4J_URI", "bolt://neo4j:7687")
    get_settings.cache_clear()
    monkeypatch.setattr(neo4jGraphPlugin, "ingestArtefact", lambda _path: {"ok": True})

    caplog.set_level("INFO")
    import anyio

    anyio.run(routes.ingest, FakeUpload())

    assert "[Ingest] Using Settings.neo4j_uri=bolt://neo4j:7687" in caplog.text
    get_settings.cache_clear()


def test_graph_health_reports_clear_neo4j_unreachable_error(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(
        neo4jGraphPlugin,
        "get_repository",
        lambda: (_ for _ in ()).throw(ServiceUnavailable("could not connect")),
    )

    with pytest.raises(HTTPException) as exc:
        routes.health()

    assert exc.value.status_code == 503
    assert exc.value.detail == "Neo4j is not reachable. Check NEO4J_URI and whether the Neo4j container is running."


def test_dynamic_cypher_identifier_sanitization():
    assert sanitize_identifier("Allocated To") == "ALLOCATED_TO"
    assert sanitize_identifier("allocated-to") == "ALLOCATED_TO"
    with pytest.raises(ValueError):
        sanitize_identifier("Requirement`) DETACH DELETE n //")


def test_generate_impact_analysis_sends_graph_context_to_llm():
    client = FakeOpenAIClient()
    graph = _fake_backup_graph()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=graph,
        ),
        settings=Settings(llm_provider="openai", openai_api_key="test-key", openai_model="test-model"),
        client=client,
    )

    assert result.summary == (
        "AI reviewed REQ-008 — Battery backup duration for impacts involving Power, Backup test, Battery risk. "
        "Backup requirement change may affect power, test, and risk artifacts."
    )
    assert result.provider == "openai"
    assert client.last_kwargs["model"] == "test-model"
    user_payload = client.user_payload
    assert user_payload["selectedRequirement"]["id"] == "REQ-008"
    assert {node["id"] for node in user_payload["impactedNodes"]} == {node.id for node in graph.nodes}
    assert {edge["source"] for edge in user_payload["impactedRelationships"]} == {edge.source for edge in graph.edges}
    assert all("score" not in node for node in user_payload["impactedNodes"])


def test_generate_impact_analysis_rejects_empty_graph():
    with pytest.raises(ValueError):
        generateImpactAnalysis(
            LLMImpactAnalysisInput(
                changeText="Any change",
                selectedRequirement=SelectedRequirement(id="REQ-001", name="Empty"),
                impactGraph=ImpactGraph(),
            ),
            settings=Settings(openai_api_key="test-key"),
            client=FakeOpenAIClient(),
        )


def test_openai_provider_missing_key_uses_rule_based_fallback():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Any change",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="openai", openai_api_key=None),
    )

    assert result.provider == "fallback"


def test_auto_provider_uses_rule_based_fallback_without_openai_or_ollama():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Any change",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="auto", openai_api_key=None, ollama_base_url="http://127.0.0.1:1"),
    )

    assert result.provider == "fallback"
    assert result.summary
    assert result.suggestedNextSteps


def test_missing_provider_defaults_to_ollama_and_does_not_call_openai():
    openai = FakeOpenAIClient(raise_error=True)
    ollama = FakeOllamaClient()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="", openai_api_key="test-key"),
        client=openai,
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert openai.last_kwargs is None
    assert len(ollama.calls) == 1


def test_generate_impact_analysis_filters_invented_nodes_from_llm_output():
    client = FakeOpenAIClient(include_invented_nodes=True)

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="openai", openai_api_key="test-key"),
        client=client,
    )

    assert all(summary.nodeId != "REQ-DOES-NOT-EXIST" for summary in result.affectedNodeSummary)
    assert all("REQ-DOES-NOT-EXIST" not in effect.affectedNodes for effect in result.rippleEffects)
    assert result.humanInTheLoopNotice == HITL_NOTICE


def test_generate_impact_analysis_handles_invalid_openai_json():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="openai", openai_api_key="test-key"),
        client=FakeOpenAIClient(invalid_json=True),
    )

    assert result.provider == "fallback"
    assert "openai returned invalid JSON" in " ".join(result.assumptionsAndLimitations)


def test_generate_impact_analysis_logs_safe_error_details(caplog):
    with caplog.at_level("ERROR"):
        result = generateImpactAnalysis(
            LLMImpactAnalysisInput(
                changeText="Adjust backup power duration",
                selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
                impactGraph=_fake_backup_graph(),
            ),
            settings=Settings(llm_provider="openai", openai_api_key="test-key"),
            client=FakeOpenAIClient(raise_error=True),
        )

    assert result.provider == "fallback"
    assert "[LLM] OpenAI request failed" in caplog.text
    assert "test-key" not in caplog.text


def test_ollama_provider_does_not_require_openai_api_key():
    ollama = FakeOllamaClient()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None, ollama_model="llama-test"),
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert ollama.calls[0]["model"] == "llama-test"
    assert ollama.calls[0]["options"]["temperature"] == 0
    assert ollama.calls[0]["options"]["num_predict"] == 900
    assert ollama.calls


def test_ollama_provider_does_not_call_openai():
    openai = FakeOpenAIClient(raise_error=True)
    ollama = FakeOllamaClient()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key="test-key"),
        client=openai,
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert openai.last_kwargs is None


def test_auto_falls_back_from_openai_quota_error_to_ollama():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="auto", openai_api_key="test-key"),
        client=FakeOpenAIClient(raise_error=True),
        ollama_client=FakeOllamaClient(),
    )

    assert result.provider == "ollama"


def test_auto_without_openai_key_tries_ollama():
    ollama = FakeOllamaClient()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="auto", openai_api_key=None),
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert len(ollama.calls) == 1


def test_ollama_failure_falls_back_to_rule_based_summary():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(raise_error=True),
    )

    assert result.provider == "fallback"
    assert "This fallback analysis is rule-based" in result.assumptionsAndLimitations[0]


def test_ollama_invalid_json_is_retried_once():
    ollama = FakeOllamaClient(invalid_first_response=True)

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert len(ollama.calls) == 2


def test_ollama_uses_compacted_graph_context():
    graph = _large_graph()
    ollama = FakeOllamaClient()

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=graph,
        ),
        settings=Settings(
            llm_provider="ollama",
            openai_api_key=None,
            ollama_max_context_nodes=5,
            ollama_max_context_edges=4,
        ),
        ollama_client=ollama,
    )

    user_payload = ollama.user_payload
    assert len(user_payload["impactedNodes"]) == 5
    assert len(user_payload["impactedRelationships"]) <= 4
    assert user_payload["impactFacts"]["connectedTestCases"]
    assert user_payload["coverageFlags"]["hasVerificationImpact"] is True
    assert user_payload["contextStats"]["truncated"] is True
    assert "compacted graph context" in " ".join(result.assumptionsAndLimitations)


def test_impact_facts_groups_nodes_by_type():
    graph = _full_category_graph()
    facts = build_impact_facts(
        SelectedRequirement(id="REQ-008", name="Battery backup duration", criticality="High"),
        graph,
    )

    assert facts["selectedRequirement"] == {
        "id": "REQ-008",
        "name": "Battery backup duration",
        "criticality": "High",
    }
    assert [node["id"] for node in facts["connectedRequirements"]] == ["REQ-009"]
    assert [node["id"] for node in facts["connectedSubsystems"]] == ["SUB-001"]
    assert [node["id"] for node in facts["connectedSoftwareModules"]] == ["SW-001"]
    assert [node["id"] for node in facts["connectedTestCases"]] == ["TEST-004"]
    assert [node["id"] for node in facts["connectedRisks"]] == ["RISK-002"]
    assert [node["id"] for node in facts["connectedIssues"]] == ["ISSUE-001"]
    assert [node["id"] for node in facts["connectedDocuments"]] == ["DOC-001"]
    assert [node["id"] for node in facts["connectedTeams"]] == ["TEAM-POWER"]
    assert facts["relationships"]


def test_coverage_flags_reflect_non_empty_categories():
    facts = build_impact_facts(
        SelectedRequirement(id="REQ-008", name="Battery backup duration"),
        _full_category_graph(),
    )

    assert build_coverage_flags(facts) == {
        "hasSubsystemImpact": True,
        "hasSoftwareImpact": True,
        "hasVerificationImpact": True,
        "hasRiskImpact": True,
        "hasIssueImpact": True,
        "hasDocumentImpact": True,
        "hasRequirementRipple": True,
    }


def test_summary_always_starts_with_ai_reviewed_requirement():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(),
    )

    assert result.summary.startswith("AI reviewed REQ-008 — Battery backup duration")


def test_non_empty_test_case_category_is_always_mentioned():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    public_text = _public_result_text(result)
    assert "test" in public_text or "verification" in public_text


def test_non_empty_risk_category_is_always_mentioned():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    assert "risk" in _public_result_text(result)


def test_non_empty_issue_category_is_always_mentioned():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    public_text = _public_result_text(result)
    assert "issue" in public_text or "anomaly" in public_text


def test_non_empty_document_category_is_always_mentioned():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    public_text = _public_result_text(result)
    assert "document" in public_text or "specification" in public_text


def test_provider_and_model_words_are_removed_from_public_output():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(provider_words=True),
    )

    public_text = _public_result_text(result)
    assert "ollama" not in public_text
    assert "openai" not in public_text
    assert "provider" not in public_text
    assert "model" not in public_text
    assert "llama" not in public_text


def test_missing_category_in_llm_output_is_patched_deterministically():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    assert result.suggestedNextSteps[:5] == [
        "Review the selected requirement and confirm the intended change.",
        "Review linked system and software areas: Power, Power manager.",
        "Review linked verification test cases: Backup test.",
        "Review linked risks and issues: Battery risk, Battery issue.",
        "Review linked documents and specifications: Power specification.",
    ]


def test_backend_preserves_three_valid_ripple_effects():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(ripple_count=3),
    )

    assert [effect.area for effect in result.rippleEffects] == [
        "Subsystem impact",
        "Verification impact",
        "Risk and document impact",
    ]


def test_backend_preserves_two_valid_ripple_effects():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(ripple_count=2),
    )

    assert len(result.rippleEffects) == 2
    assert [effect.area for effect in result.rippleEffects] == ["Subsystem impact", "Verification impact"]


def test_backend_trims_ripple_effects_to_three_not_one():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(ripple_count=4),
    )

    assert len(result.rippleEffects) == 3
    assert result.rippleEffects[0].area == "Subsystem impact"
    assert result.rippleEffects[2].area == "Risk and document impact"


def test_backend_creates_graph_specific_fallback_ripple_effects_when_llm_returns_none():
    result = _analysis_with_sparse_llm_output(_full_category_graph())

    assert 1 <= len(result.rippleEffects) <= 3
    assert result.rippleEffects[0].area == "Power"
    assert all(effect.area != "Knowledge graph impact" for effect in result.rippleEffects)


def test_generic_ripple_effect_triggers_retry_and_preserves_specific_retry_output():
    ollama = FakeOllamaClient(generic_first_response=True, ripple_count=3)

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=ollama,
    )

    assert len(ollama.calls) == 2
    assert [effect.area for effect in result.rippleEffects] == [
        "Subsystem impact",
        "Verification impact",
        "Risk and document impact",
    ]


def test_generic_ripple_effect_falls_back_to_real_graph_nodes_when_retry_is_generic():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="openai", openai_api_key="test-key"),
        client=FakeOpenAIClient(generic_ripple=True),
    )

    assert [effect.area for effect in result.rippleEffects[:3]] == [
        "Power",
        "Power manager",
        "Verification impact",
    ]
    assert all(effect.area != "Knowledge graph impact" for effect in result.rippleEffects)


def test_fallback_summary_uses_actual_graph_node_names():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_full_category_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(raise_error=True),
    )

    assert result.summary.startswith("AI reviewed REQ-008")
    assert "Power" in result.summary
    assert "Power manager" in result.summary
    assert "Backup test" in result.summary


def test_launch_window_graph_summary_mentions_actual_connected_nodes():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Tighten launch window command timing",
            selectedRequirement=SelectedRequirement(id="REQ-030", name="Launch Window Constraint"),
            impactGraph=_launch_window_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(sparse_response=True),
    )

    assert "Ground Support & Launch Control" in result.summary
    assert "Ground Command Uplink Latency" in result.summary
    assert "Countdown & Launch Sequencer" in result.summary
    assert all(effect.area != "Knowledge graph impact" for effect in result.rippleEffects)


def test_ollama_timeout_causes_shorter_retry_then_success():
    ollama = FakeOllamaClient(raise_first_response=True)

    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_large_graph(),
        ),
        settings=Settings(
            llm_provider="ollama",
            openai_api_key=None,
            ollama_max_context_nodes=20,
            ollama_max_context_edges=30,
        ),
        ollama_client=ollama,
    )

    assert result.provider == "ollama"
    assert len(ollama.calls) == 2
    retry_payload = ollama.user_payloads[1]
    assert retry_payload["contextStats"]["nodesSent"] <= 10
    assert retry_payload["contextStats"]["edgesSent"] <= 15


def test_ollama_retry_failure_preserves_fallback_reason():
    result = generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=_fake_backup_graph(),
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(raise_error=True),
    )

    assert result.provider == "fallback"
    assert "Ollama unavailable" in " ".join(result.assumptionsAndLimitations)


def test_llm_health_reports_ollama_model_available(monkeypatch):
    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def read(self):
            return b'{"models":[{"name":"llama3.2:3b"}]}'

    monkeypatch.setattr(impactAnalysisLLM.urllib.request, "urlopen", lambda *_args, **_kwargs: FakeResponse())

    health = impactAnalysisLLM.get_llm_health(
        Settings(llm_provider="ollama", openai_api_key=None, ollama_model="llama3.2:3b")
    )

    assert health["provider"] == "ollama"
    assert health["ollama"]["reachable"] is True
    assert health["ollama"]["modelAvailable"] is True
    assert health["ollama"]["tagsLatencyMs"] is not None
    assert health["config"]["timeoutMs"] == 300000
    assert health["config"]["maxContextNodes"] == 20
    assert health["config"]["maxContextEdges"] == 30
    assert health["config"]["maxDescriptionChars"] == 180
    assert health["openai"]["configured"] is False


def test_llm_test_endpoint_returns_generation_result(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(
        routes,
        "test_ollama_generation",
        lambda prompt: {
            "provider": "ollama",
            "success": True,
            "durationMs": 12,
            "rawResponsePreview": '{"summary":"hello"}',
            "parsed": {"summary": "hello"},
        },
    )
    client = TestClient(app)

    response = client.post("/api/graph/llm-test", json={"prompt": "Return JSON only with summary saying hello."})

    assert response.status_code == 200
    assert response.json()["success"] is True


def test_llm_endpoint_returns_service_unavailable_without_api_key(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(routes, "generateImpactAnalysis", lambda _input: (_ for _ in ()).throw(ImpactAnalysisUnavailable("OPENAI_API_KEY is not configured.")))
    client = TestClient(app)

    response = client.post(
        "/api/graph/impact-analysis",
        json={
            "changeText": "Adjust backup power duration",
            "selectedRequirement": {"id": "REQ-008", "type": "Requirement", "name": "Battery backup duration"},
            "impactGraph": _fake_backup_graph().model_dump(),
        },
    )

    assert response.status_code == 503
    assert response.json()["detail"]["error"] == "AI_ANALYSIS_UNAVAILABLE"


def test_llm_endpoint_returns_structured_failure(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(routes, "generateImpactAnalysis", lambda _input: (_ for _ in ()).throw(ImpactAnalysisFailed("OpenAI request failed: quota exceeded")))
    client = TestClient(app)

    response = client.post(
        "/api/graph/impact-analysis",
        json={
            "changeText": "Adjust backup power duration",
            "selectedRequirement": {"id": "REQ-008", "type": "Requirement", "name": "Battery backup duration"},
            "impactGraph": _fake_backup_graph().model_dump(),
        },
    )

    assert response.status_code == 502
    assert response.json()["detail"] == {
        "error": "AI_ANALYSIS_FAILED",
        "message": "AI analysis could not be generated for this graph. Please review the impact map manually.",
    }


def test_llm_endpoint_normal_response_hides_provider_and_debug_details(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(
        routes,
        "generateImpactAnalysis",
        lambda _input: _fake_llm_result(provider="ollama"),
    )
    client = TestClient(app)

    response = client.post(
        "/api/graph/impact-analysis",
        json={
            "changeText": "Adjust backup power duration",
            "selectedRequirement": {"id": "REQ-008", "type": "Requirement", "name": "Battery backup duration"},
            "impactGraph": _fake_backup_graph().model_dump(),
        },
    )

    payload = response.json()
    assert response.status_code == 200
    assert set(payload["analysis"]) == {"summary", "rippleEffects", "suggestedNextSteps"}
    assert set(payload["analysis"]["rippleEffects"][0]) == {"area", "explanation"}
    serialized = str(payload)
    assert "provider" not in serialized
    assert "engineeringReviewChecklist" not in serialized
    assert "assumptionsAndLimitations" not in serialized
    assert "humanInTheLoopNotice" not in serialized
    assert "ollama" not in serialized.lower()


def test_llm_endpoint_debug_response_includes_provider_diagnostics(monkeypatch):
    from app.plugins.neo4jGraphPlugin import routes

    monkeypatch.setattr(
        routes,
        "generateImpactAnalysis",
        lambda _input: _fake_llm_result(provider="fallback"),
    )
    client = TestClient(app)

    response = client.post(
        "/api/graph/impact-analysis",
        json={
            "changeText": "Adjust backup power duration",
            "selectedRequirement": {"id": "REQ-008", "type": "Requirement", "name": "Battery backup duration"},
            "impactGraph": _fake_backup_graph().model_dump(),
            "debug": True,
        },
    )

    assert response.status_code == 200
    assert response.json()["debug"]["provider"] == "fallback"
    assert response.json()["debug"]["fallbackUsed"] is True


class FakeRepository:
    def search_requirement(self, change_text: str):
        lower_text = change_text.lower()
        if "nonsense" in lower_text:
            return []
        if "flight control" in lower_text:
            return [
                RequirementCandidate(
                    id="REQ-004",
                    name="Flight Control Response Time",
                    description="The flight control response shall stay within the configured time limit.",
                    criticality="Critical",
                    score=11.0,
                )
            ]
        if "cyber" in lower_text:
            return [
                RequirementCandidate(
                    id="REQ-020",
                    name="Cybersecurity Zoning",
                    description="Passenger and avionics domains shall be isolated.",
                    criticality="High",
                    score=10.0,
                )
            ]
        return [
            RequirementCandidate(
                id="REQ-008",
                name="Battery backup duration",
                description="The battery backup shall last 45 minutes.",
                criticality="High",
                score=13.0,
            ),
            RequirementCandidate(
                id="REQ-002",
                name="Display brightness",
                description="The display shall be readable in sunlight.",
                criticality="Medium",
                score=1.0,
            ),
        ]

    def get_impact_graph(self, requirement_id: str, depth: int = 2):
        if requirement_id == "REQ-004":
            return ImpactGraph(
                nodes=[
                    ImpactNode(id="REQ-004", label="REQ-004", type="Requirement", name="Flight Control Response Time", status="selected", hop=0),
                    ImpactNode(id="SW-004", label="SW-004", type="SoftwareModule", name="Flight control loop", status="impacted", hop=1),
                    ImpactNode(id="TC-004", label="TC-004", type="TestCase", name="Control latency test", status="impacted", hop=1),
                ],
                edges=[
                    ImpactEdge(id="REQ-004::TRACED_TO::SW-004", source="REQ-004", target="SW-004", relationship="TRACED_TO", status="ontology-link", hop=1),
                    ImpactEdge(id="REQ-004::VERIFIED_BY::TC-004", source="REQ-004", target="TC-004", relationship="VERIFIED_BY", status="ontology-link", hop=1),
                ],
            )
        if requirement_id == "REQ-020":
            return ImpactGraph(
                nodes=[
                    ImpactNode(id="REQ-020", label="REQ-020", type="Requirement", name="Cybersecurity Zoning", status="selected", hop=0),
                    ImpactNode(id="SUB-020", label="SUB-020", type="Subsystem", name="Network Security", status="impacted", hop=1),
                ],
                edges=[
                    ImpactEdge(id="REQ-020::ALLOCATED_TO::SUB-020", source="REQ-020", target="SUB-020", relationship="ALLOCATED_TO", status="ontology-link", hop=1),
                ],
            )
        return ImpactGraph(
            nodes=[
                ImpactNode(
                    id=requirement_id,
                    label=requirement_id,
                    type="Requirement",
                    name="Battery backup duration",
                    status="selected",
                    hop=0,
                ),
                ImpactNode(id="SUB-001", label="SUB-001", type="Subsystem", name="Power", status="impacted", hop=1),
                ImpactNode(id="TEST-004", label="TEST-004", type="Test", name="Backup test", status="impacted", hop=1),
                ImpactNode(id="TEAM-POWER", label="TEAM-POWER", type="Team", name="Power team", status="impacted", hop=2),
                ImpactNode(id="RISK-002", label="RISK-002", type="Risk", name="Battery risk", status="impacted", hop=1),
            ],
            edges=[
                ImpactEdge(
                    id="REQ-008::ALLOCATED_TO::SUB-001",
                    source="REQ-008",
                    target="SUB-001",
                    relationship="ALLOCATED_TO",
                    status="ontology-link",
                    hop=1,
                ),
                ImpactEdge(
                    id="TEST-004::VALIDATES::REQ-008",
                    source="TEST-004",
                    target="REQ-008",
                    relationship="VALIDATES",
                    status="ontology-link",
                    hop=1,
                ),
            ],
        )


def _sample_workbook(tmp_path: Path) -> Path:
    workbook_path = tmp_path / "invalid-edge.xlsx"
    workbook = Workbook()
    nodes = workbook.active
    nodes.title = "nodes"
    nodes.append(["id", "type", "name", "description", "criticality"])
    nodes.append(["REQ-008", "Requirement", "Battery backup", "Battery backup duration", "High"])
    nodes.append(["SUB-001", "Subsystem", "Power subsystem", "UPS and battery pack", "Medium"])
    nodes.append(["TEAM-POWER", "Team", "Power team", "Owner team", "Low"])
    edges = workbook.create_sheet("edges")
    edges.append(["source_id", "target_id", "relationship", "description"])
    edges.append(["REQ-008", "SUB-001", "ALLOCATED_TO", "Allocated to power"])
    edges.append(["REQ-008", "TEAM-POWER", "OWNED_BY", "Invalid for this ontology"])
    ontology = workbook.create_sheet("ontology")
    ontology.append(["source_entity", "relationship", "target_entity"])
    ontology.append(["Requirement", "ALLOCATED_TO", "Subsystem"])
    workbook.save(workbook_path)
    return workbook_path


def _fake_backup_graph() -> ImpactGraph:
    return ImpactGraph(
        nodes=[
            ImpactNode(id="REQ-008", label="REQ-008", type="Requirement", name="Battery backup duration", status="selected", hop=0),
            ImpactNode(id="SUB-001", label="SUB-001", type="Subsystem", name="Power", status="impacted", hop=1),
            ImpactNode(id="TEST-004", label="TEST-004", type="TestCase", name="Backup test", status="impacted", hop=1),
            ImpactNode(id="RISK-002", label="RISK-002", type="Risk", name="Battery risk", status="impacted", hop=1),
        ],
        edges=[
            ImpactEdge(id="REQ-008::ALLOCATED_TO::SUB-001", source="REQ-008", target="SUB-001", relationship="ALLOCATED_TO", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-008::VERIFIED_BY::TEST-004", source="REQ-008", target="TEST-004", relationship="VERIFIED_BY", status="ontology-link", hop=1),
        ],
    )


def _full_category_graph() -> ImpactGraph:
    return ImpactGraph(
        nodes=[
            ImpactNode(id="REQ-008", label="REQ-008", type="Requirement", name="Battery backup duration", status="selected", hop=0),
            ImpactNode(id="REQ-009", label="REQ-009", type="Requirement", name="Charging behavior", status="impacted", hop=2),
            ImpactNode(id="SUB-001", label="SUB-001", type="Subsystem", name="Power", status="impacted", hop=1),
            ImpactNode(id="SW-001", label="SW-001", type="SoftwareModule", name="Power manager", status="impacted", hop=1),
            ImpactNode(id="TEST-004", label="TEST-004", type="TestCase", name="Backup test", status="impacted", hop=1),
            ImpactNode(id="RISK-002", label="RISK-002", type="Risk", name="Battery risk", status="impacted", hop=1),
            ImpactNode(id="ISSUE-001", label="ISSUE-001", type="Issue", name="Battery issue", status="impacted", hop=2),
            ImpactNode(id="DOC-001", label="DOC-001", type="Document", name="Power specification", status="impacted", hop=2),
            ImpactNode(id="TEAM-POWER", label="TEAM-POWER", type="Team", name="Power team", status="impacted", hop=2),
        ],
        edges=[
            ImpactEdge(id="REQ-008::ALLOCATED_TO::SUB-001", source="REQ-008", target="SUB-001", relationship="ALLOCATED_TO", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-008::TRACED_TO::SW-001", source="REQ-008", target="SW-001", relationship="TRACED_TO", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-008::VERIFIED_BY::TEST-004", source="REQ-008", target="TEST-004", relationship="VERIFIED_BY", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-008::HAS_RISK::RISK-002", source="REQ-008", target="RISK-002", relationship="HAS_RISK", status="ontology-link", hop=1),
            ImpactEdge(id="RISK-002::RAISED_AS::ISSUE-001", source="RISK-002", target="ISSUE-001", relationship="RAISED_AS", status="ontology-link", hop=2),
            ImpactEdge(id="REQ-008::SPECIFIED_IN::DOC-001", source="REQ-008", target="DOC-001", relationship="SPECIFIED_IN", status="ontology-link", hop=2),
            ImpactEdge(id="SUB-001::OWNED_BY::TEAM-POWER", source="SUB-001", target="TEAM-POWER", relationship="OWNED_BY", status="ontology-link", hop=2),
            ImpactEdge(id="REQ-008::RELATED_TO::REQ-009", source="REQ-008", target="REQ-009", relationship="RELATED_TO", status="ontology-link", hop=2),
        ],
    )


def _launch_window_graph() -> ImpactGraph:
    return ImpactGraph(
        nodes=[
            ImpactNode(id="REQ-030", label="REQ-030", type="Requirement", name="Launch Window Constraint", status="selected", hop=0),
            ImpactNode(id="SUB-030", label="SUB-030", type="Subsystem", name="Ground Support & Launch Control", status="impacted", hop=1),
            ImpactNode(id="SW-030", label="SW-030", type="SoftwareModule", name="Countdown & Launch Sequencer", status="impacted", hop=1),
            ImpactNode(id="TEST-030", label="TEST-030", type="TestCase", name="Ground Command Uplink Latency", status="impacted", hop=1),
        ],
        edges=[
            ImpactEdge(id="REQ-030::ALLOCATED_TO::SUB-030", source="REQ-030", target="SUB-030", relationship="ALLOCATED_TO", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-030::TRACED_TO::SW-030", source="REQ-030", target="SW-030", relationship="TRACED_TO", status="ontology-link", hop=1),
            ImpactEdge(id="REQ-030::VERIFIED_BY::TEST-030", source="REQ-030", target="TEST-030", relationship="VERIFIED_BY", status="ontology-link", hop=1),
        ],
    )


def _analysis_with_sparse_llm_output(graph: ImpactGraph):
    return generateImpactAnalysis(
        LLMImpactAnalysisInput(
            changeText="Adjust backup power duration",
            selectedRequirement=SelectedRequirement(id="REQ-008", name="Battery backup duration"),
            impactGraph=graph,
        ),
        settings=Settings(llm_provider="ollama", openai_api_key=None),
        ollama_client=FakeOllamaClient(sparse_response=True),
    )


def _public_result_text(result) -> str:
    return " ".join(
        [
            result.summary,
            *[effect.area for effect in result.rippleEffects],
            *[effect.explanation for effect in result.rippleEffects],
            *result.suggestedNextSteps,
        ]
    ).lower()


def _fake_llm_result(provider: str = "ollama"):
    from app.plugins.neo4jGraphPlugin.types import LLMImpactAnalysisResult, RippleEffect

    return LLMImpactAnalysisResult(
        provider=provider,
        summary="Concise summary.",
        rippleEffects=[RippleEffect(area="Verification", explanation="Review linked tests.")],
        suggestedNextSteps=["Review the selected requirement."],
        engineeringReviewChecklist=["Hidden checklist item."],
        assumptionsAndLimitations=["Fallback reason: hidden technical detail"],
        humanInTheLoopNotice=HITL_NOTICE,
    )


def _large_graph() -> ImpactGraph:
    nodes = [
        ImpactNode(id="REQ-008", label="REQ-008", type="Requirement", name="Battery backup duration", status="selected", hop=0)
    ]
    edges = []
    for index in range(1, 35):
        node_type = "Subsystem" if index % 3 == 0 else "TestCase" if index % 3 == 1 else "Risk"
        node_id = f"NODE-{index:03d}"
        nodes.append(
            ImpactNode(
                id=node_id,
                label=node_id,
                type=node_type,
                name=f"Node {index}",
                criticality="Critical" if index % 5 == 0 else "High" if index % 4 == 0 else "Medium",
                status="impacted",
                hop=(index % 3) + 1,
            )
        )
        edges.append(
            ImpactEdge(
                id=f"REQ-008::RELATES_TO::{node_id}",
                source="REQ-008",
                target=node_id,
                relationship="RELATES_TO",
                status="ontology-link",
                hop=(index % 3) + 1,
            )
        )
    return ImpactGraph(nodes=nodes, edges=edges)


class FakeOpenAIClient:
    def __init__(
        self,
        include_invented_nodes: bool = False,
        invalid_json: bool = False,
        raise_error: bool = False,
        generic_ripple: bool = False,
    ):
        self.chat = self
        self.completions = self
        self.include_invented_nodes = include_invented_nodes
        self.invalid_json = invalid_json
        self.raise_error = raise_error
        self.generic_ripple = generic_ripple
        self.last_kwargs = None
        self.user_payload = None

    def create(self, **kwargs):
        import json
        from types import SimpleNamespace

        self.last_kwargs = kwargs
        self.user_payload = json.loads(kwargs["messages"][1]["content"])
        if self.raise_error:
            raise RuntimeError("OpenAI transport failed")
        if self.invalid_json:
            content = "not json"
            return SimpleNamespace(
                choices=[
                    SimpleNamespace(
                        message=SimpleNamespace(content=content)
                    )
                ]
            )
        affected_nodes = [
            {
                "nodeId": "SUB-001",
                "nodeName": "Power",
                "nodeType": "Subsystem",
                "whyItMatters": "Power allocation may need engineering review.",
            }
        ]
        ripple_nodes = ["SUB-001", "TEST-004"]
        if self.include_invented_nodes:
            affected_nodes.append(
                {
                    "nodeId": "REQ-DOES-NOT-EXIST",
                    "nodeName": "Invented",
                    "nodeType": "Requirement",
                    "whyItMatters": "Should be filtered.",
                }
            )
            ripple_nodes.append("REQ-DOES-NOT-EXIST")
        ripple_effects = [
            {
                "area": "Power and verification",
                "explanation": "Direct graph links show subsystem and test impact.",
                "affectedNodes": ripple_nodes,
            }
        ]
        if self.generic_ripple:
            ripple_effects = [
                {
                    "area": "Knowledge graph impact",
                    "explanation": "Review the connected impacted nodes and relationships before approving the change.",
                    "affectedNodes": ripple_nodes,
                }
            ]
        content = json.dumps(
            {
                "summary": "Backup requirement change may affect power, test, and risk artifacts.",
                "rippleEffects": ripple_effects,
                "affectedNodeSummary": affected_nodes,
                "suggestedNextSteps": ["Review impacted subsystem and test owners."],
                "engineeringReviewChecklist": ["Confirm graph links with responsible engineer."],
                "assumptionsAndLimitations": [],
                "humanInTheLoopNotice": "" if self.include_invented_nodes else HITL_NOTICE,
            }
        )
        return SimpleNamespace(
            choices=[
                SimpleNamespace(
                    message=SimpleNamespace(content=content)
                )
            ]
        )


class FakeOllamaClient:
    def __init__(
        self,
        invalid_first_response: bool = False,
        raise_error: bool = False,
        raise_first_response: bool = False,
        sparse_response: bool = False,
        provider_words: bool = False,
        ripple_count: int = 1,
        generic_first_response: bool = False,
    ):
        self.invalid_first_response = invalid_first_response
        self.raise_error = raise_error
        self.raise_first_response = raise_first_response
        self.sparse_response = sparse_response
        self.provider_words = provider_words
        self.ripple_count = ripple_count
        self.generic_first_response = generic_first_response
        self.calls = []
        self.user_payloads = []
        self.user_payload = None

    def __call__(self, payload):
        import json

        self.calls.append(payload)
        self.user_payload = json.loads(payload["messages"][1]["content"])
        self.user_payloads.append(self.user_payload)
        if self.raise_error:
            raise RuntimeError("Ollama unavailable")
        if self.raise_first_response and len(self.calls) == 1:
            raise TimeoutError("timed out")
        if self.invalid_first_response and len(self.calls) == 1:
            return {"message": {"content": "not json"}}
        if self.generic_first_response and len(self.calls) == 1:
            return {
                "message": {
                    "content": json.dumps(
                        {
                            "summary": "Generic graph review.",
                            "rippleEffects": [
                                {
                                    "area": "Knowledge graph impact",
                                    "explanation": "Review the connected impacted nodes and relationships before approving the change.",
                                }
                            ],
                            "suggestedNextSteps": ["Review the graph."],
                        }
                    )
                }
            }
        if self.sparse_response:
            return {
                "message": {
                    "content": json.dumps(
                        {
                            "summary": "Graph-connected engineering review is needed.",
                            "rippleEffects": [],
                            "suggestedNextSteps": [],
                        }
                    )
                }
            }
        if self.provider_words:
            return {
                "message": {
                    "content": json.dumps(
                        {
                            "summary": "Ollama reviewed REQ-008 using llama3.2 model.",
                            "rippleEffects": [
                                {
                                    "area": "OpenAI provider wording",
                                    "explanation": "Fallback provider and Ollama model words should be removed.",
                                }
                            ],
                            "suggestedNextSteps": ["Do not expose provider or model details."],
                        }
                    )
                }
            }
        if self.ripple_count != 1:
            ripple_effects = [
                {
                    "area": "Subsystem impact",
                    "explanation": "Subsystem allocation should be reviewed for the requirement change.",
                },
                {
                    "area": "Verification impact",
                    "explanation": "Linked verification tests should be checked before approval.",
                },
                {
                    "area": "Risk and document impact",
                    "explanation": "Risk and specification links may need coordinated updates.",
                },
                {
                    "area": "Team coordination",
                    "explanation": "Responsible teams should confirm ownership of the change.",
                },
            ]
            return {
                "message": {
                    "content": json.dumps(
                        {
                            "summary": "Graph-connected engineering impacts were identified.",
                            "rippleEffects": ripple_effects[: self.ripple_count],
                            "suggestedNextSteps": ["Review the selected requirement."],
                        }
                    )
                }
            }
        return {
            "message": {
                "content": json.dumps(
                    {
                        "summary": "Local analysis identified graph-connected requirement impacts.",
                        "rippleEffects": [
                            {
                                "area": "Graph analysis",
                                "explanation": "The analysis used only returned graph nodes.",
                                "affectedNodes": ["SUB-001", "TEST-004"],
                            }
                        ],
                        "suggestedNextSteps": ["Review locally generated suggestions."],
                        "engineeringReviewChecklist": ["Confirm impacted nodes with an engineer."],
                        "assumptionsAndLimitations": ["Generated from the provided impact graph."],
                        "humanInTheLoopNotice": HITL_NOTICE,
                    }
                )
            }
        }
