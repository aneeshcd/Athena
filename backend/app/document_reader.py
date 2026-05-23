from io import BytesIO
from pathlib import Path
from fastapi import UploadFile
from openpyxl import load_workbook
from pypdf import PdfReader


async def extract_upload_text(file: UploadFile) -> str:
    content = await file.read()
    suffix = Path(file.filename or "").suffix.lower()

    if suffix == ".pdf":
        return _extract_pdf(content)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_excel(content)
    if suffix in {".txt", ".md", ".csv", ".tsv", ".json"}:
        return content.decode("utf-8", errors="replace")

    return content.decode("utf-8", errors="replace")


def _extract_pdf(content: bytes) -> str:
    reader = PdfReader(BytesIO(content))
    pages: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(f"[page {index}]\n{text}")
    return "\n\n".join(pages)


def _extract_excel(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[sheet {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                lines.append("\t".join(values))
    return "\n".join(lines)
