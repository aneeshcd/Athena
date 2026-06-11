from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.plugins.neo4jGraphPlugin.ontologyValidator import canonical_entity_type, canonical_relationship
from app.plugins.neo4jGraphPlugin.types import ArtefactEdge, ArtefactNode, OntologyRule, ParsedArtefact


REQUIRED_SHEETS = {"nodes", "edges", "ontology"}


def read_excel_artefact(file_path: str) -> ParsedArtefact:
    workbook = load_workbook(Path(file_path), read_only=True, data_only=True)
    try:
        sheet_by_name = {name.lower(): name for name in workbook.sheetnames}
        missing = sorted(REQUIRED_SHEETS - set(sheet_by_name))
        if missing:
            raise ValueError(f"Workbook is missing required sheet(s): {', '.join(missing)}")

        nodes = [
            ArtefactNode(
                id=_required(row, "id"),
                type=canonical_entity_type(_required(row, "type")),
                name=_required(row, "name"),
                description=_optional(row, "description"),
                criticality=_optional(row, "criticality"),
            )
            for row in _rows(workbook[sheet_by_name["nodes"]])
        ]
        edges = [
            ArtefactEdge(
                source_id=_required(row, "source_id"),
                target_id=_required(row, "target_id"),
                relationship=canonical_relationship(_required(row, "relationship")),
                description=_optional(row, "description"),
            )
            for row in _rows(workbook[sheet_by_name["edges"]])
        ]
        ontology = [
            OntologyRule(
                source_entity=canonical_entity_type(_required(row, "source_entity")),
                relationship=canonical_relationship(_required(row, "relationship")),
                target_entity=canonical_entity_type(_required(row, "target_entity")),
            )
            for row in _rows(workbook[sheet_by_name["ontology"]])
        ]
        return ParsedArtefact(nodes=nodes, edges=edges, ontology=ontology)
    finally:
        workbook.close()


def _rows(sheet) -> list[dict[str, Any]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []
    normalized_headers = [str(header or "").strip().lower() for header in headers]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        row = {header: value for header, value in zip(normalized_headers, values, strict=False)}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def _required(row: dict[str, Any], key: str) -> str:
    value = _optional(row, key)
    if not value:
        raise ValueError(f"Required column {key!r} is empty.")
    return value


def _optional(row: dict[str, Any], key: str) -> str:
    value = row.get(key)
    return "" if value is None else str(value).strip()
